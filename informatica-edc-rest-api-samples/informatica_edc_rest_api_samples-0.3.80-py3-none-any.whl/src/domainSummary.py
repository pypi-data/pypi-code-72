"""
Created on Jan 7, 2020

@author: dwrigley

usage:
  domainSummary.py -h to see command-line options

  output written to domain_summary.csv

Note:  requires python 3 (3.6+)

get a list of all custom attributes in EDC & count the usage of each attribute
the id would be used for any search/custom import activiies
output printed to the console

Informatica tested with:  v10.4.0, 10.2.2sp1
Jac. tested with: v10.4.1
"""
import argparse
import csv
import os
import sys
import time
from pathlib import PurePath

import requests
import urllib3
from openpyxl import Workbook

from src.edc_utilities.edcSessionHelper import EDCSession

urllib3.disable_warnings()


class MemVariables:
    """
    empty class for storing variables & not making them global
    """

    edcSession: EDCSession = EDCSession()
    workbook = Workbook()


start_time = time.time()
# initialize http header - as a dict
# header = {}
# auth = None

pageSize = 5000  # number of objects for each page/chunk

# the csv lineage file to write to
csv_file_name = "domain_summary.csv"
excelFileName = "domain_summary.xlsx"
# the path to write to - can be overwritten by -o cmdlink parameter
csv_file_path = "../out/"

# edcSession = EDCSession()
parser = argparse.ArgumentParser(parents=[MemVariables.edcSession.argparser])
parser.add_argument("-o", "--output", default="./out", help="output folder - e.g. .out")
# parser.add_argument(
#     "-xls",
#     "--exceloutput",
#     dest="writeToExcel",
#     action="store_true",
#     help="write output to excel (single file)",
# )
parser.add_argument(
    "--csvoutput",
    dest="writeToCSV",
    default=False,
    action="store_true",
    help="write output to csv (multiple files) in addition to excel",
)


def search_summary_v1(session, resturl, querystring):
    try:
        resp = session.get(resturl, params=querystring, timeout=3)
        # print(f"api status code={resp.status_code}")
        return resp.status_code, resp.json()
    except requests.exceptions.RequestException as e:
        print("Error connecting to : " + resturl)
        print(e)
        # exit if we can't connect
        return 0, None


def setup_excel_workbook(workbook_folder, workbook_name):
    MemVariables.workbook = Workbook()
    # set file path
    # MemVariables.wbpath = f"{wbFolder}/{wbName}"


def getDomainQueryStats(resultJson, domainData, classData, resourceData):
    # analyze - all domains
    all_facets = resultJson["facets"]["facetFields"]
    # print(f"facet field count = {len(all_facets)}")

    for aFacet in all_facets:
        # print(f"\t reading facet name={aFacet['fieldName']}")
        facetName = aFacet["fieldName"].rsplit(".", 1)[-1]
        if aFacet["fieldName"].startswith("com.infa.ldm.profiling.dataDomains"):
            for row in aFacet["rows"]:
                # print(f"\t\t{facetName} {row['value']}={row['count']}")
                # if statement necessary for pre 10.2.2hf1
                if (row["count"]) > 0:
                    domainInst = domainData.get(row["value"], {})
                    domainInst[facetName] = domainInst.get(facetName, 0) + row["count"]
                    domainData[row["value"]] = domainInst
        if aFacet["fieldName"] == "core.classType":
            for row in aFacet["rows"]:
                # for older versions (before 10.2.2hf1)
                # facets returned 0 values (which we don't want)
                if row["count"] > 0:
                    classData[row["value"]] = row["count"]
        if aFacet["fieldName"] == "core.resource_name":
            for row in aFacet["rows"]:
                # for older versions (before 10.2.2hf1)
                # facets returned 0 values (which we don't want)
                if row["count"] > 0:
                    resourceData[row["value"]] = row["count"]

    # MemVariables objects are updated, nothing to return
    return


def printDomainSumarytoExcelWorksheet(workbook, sheetName, domainData):
    sheet = workbook.active
    if sheet.title == "Sheet":
        ws1 = workbook.worksheets[0]
    else:
        ws1 = workbook.create_sheet()
    ws1.title = sheetName
    ws1.column_dimensions["A"].width = 50
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 12
    ws1.append(("Domain Name", "All Count", "Accepted", "Inferred", "Rejected"))

    rowcount = 0
    for domain, domain_obj in sorted(domainData.items(), key=lambda x: x[0].lower()):
        rowcount += 1
        ws1.append(
            (
                domain,
                domain_obj.get("dataDomainsAll", 0),
                domain_obj.get("dataDomainsAccepted", 0),
                domain_obj.get("dataDomainsInferred", 0),
                domain_obj.get("dataDomainsRejected", 0),
            )
        )
    ws1.auto_filter.ref = f"A1:E{rowcount + 1}"
    # print(f"worksheet autofilter {sheetName} = A1:E{rowcount+1}")


def printDomainSummaryToCsv(outputFile, domainData):
    # create and initialize the header for the output csv file
    with open(outputFile, "w", newline="", encoding="utf-8") as fCSVFile:
        # print("\t\tdomain summary csv file initialized: " + outputFile)
        colWriter = csv.writer(fCSVFile)
        colWriter.writerow(
            [
                "Domain Name",
                "All Count",
                "Accepted Count",
                "Inferred Count",
                "Rejected Count",
            ]
        )

        for domain, domain_obj in sorted(domainData.items()):
            colWriter.writerow(
                [
                    domain,
                    domain_obj.get("dataDomainsAll", 0),
                    domain_obj.get("dataDomainsAccepted", 0),
                    domain_obj.get("dataDomainsInferred", 0),
                    domain_obj.get("dataDomainsRejected", 0),
                ]
            )


def main():
    """
    get a count of all data domains that are used, and the counts for all,
    accepted, inferred & rejected
    collect the list of resources, then get the same stats per resource

    write the results to an excel file, with 1 tab for overall counts,
    and 1 tab per resourceType
    2 other tabs (counts for classes & resources found):-
        classes used
        resources used
    """
    p = PurePath(sys.argv[0])
    print(f"{p.name} starting in {os.getcwd()}")

    csv_file_path = ".."
    # read any command-line args passed
    # (only needed if using extra args from what is used in edcSession
    print("\treading command-line specific to domainSummary extractor")
    args, unknown = parser.parse_known_args()
    # initialize http session to EDC, storing the baseurl
    MemVariables.edcSession.initUrlAndSessionFromEDCSettings()
    # print(
    #     f"args from cmdline/env vars: url={MemVariables.edcSession.baseUrl}"
    #     f"  session={MemVariables.edcSession.session}"
    #     f"arg={args}"
    # )

    # create the output path if it does not exist
    if args.output is not None:
        csv_file_path = args.output
        print(f"\t\toutput path={csv_file_path}")
        if csv_file_path != "":
            if not os.path.exists(csv_file_path):
                os.makedirs(csv_file_path)

    MemVariables.outputCsvFile = csv_file_path + "/" + csv_file_name
    print(f"\t\toutput excel file=True")
    print(f"\t\toutput csv files={args.writeToCSV}")
    MemVariables.outputExcelFile = csv_file_path + "/" + excelFileName

    # format the query parameters for finding all domains (+ rejected) for all resources
    querystring = {
        "q": (
            "com.infa.ldm.profiling.dataDomainsAll:* "
            "OR com.infa.ldm.profiling.dataDomainsRejected:*"
        ),
        "offset": "0",
        "pageSize": "1",
        "hl": "false",
        "related": "false",
        "rootto": "false",
        "facet.field": [
            "core.classType",
            "core.resource_name",
            "com.infa.ldm.profiling.dataDomainsInferred",
            "com.infa.ldm.profiling.dataDomainsAll",
            "com.infa.ldm.profiling.dataDomainsRejected",
            "com.infa.ldm.profiling.dataDomainsAccepted",
        ],
        # "includeRefObjects": "false",
    }

    print(f"\nexecuting search for domains in use q={querystring['q']}")
    print(f"\tusing facets: {querystring['facet.field']}")
    resturl = MemVariables.edcSession.baseUrl + "/access/1/catalog/data/search"
    rc, domainJson = search_summary_v1(MemVariables.edcSession.session, resturl, querystring)
    print(f"query rc= {rc}")
    if rc != 200:
        print(f"error running query: {rc} {domainJson}")
        print("exiting")
        return

    itemCount = domainJson["totalCount"]
    print(f"items found={domainJson['totalCount']:,}")

    MemVariables.domainData = dict()
    MemVariables.classData = dict()
    MemVariables.resourceData = dict()

    # analyze - all domains (including rejected)
    getDomainQueryStats(domainJson, MemVariables.domainData, MemVariables.classData, MemVariables.resourceData)

    # print to csv file
    print(f"domains used: {len(MemVariables.domainData)}")
    if args.writeToCSV:
        printDomainSummaryToCsv(f"out/domainSummary.csv", MemVariables.domainData)
    # if args.writeToExcel:
    setup_excel_workbook(csv_file_path, "datadomain_summary.xlsx")
    printDomainSumarytoExcelWorksheet(MemVariables.workbook, "all domains", MemVariables.domainData)

    print(f"\nclasses using any domains {len(MemVariables.classData)}")
    ws1 = MemVariables.workbook.create_sheet()
    ws1.title = "classes used"
    ws1.append(("Class Type", "Count"))
    ws1.column_dimensions["A"].width = 50
    for theclass, countdomains in sorted(MemVariables.classData.items()):
        print(f"\tclass {theclass}:{countdomains}")
        ws1.append((theclass, countdomains))

    print(f"\nresources using any domains {len(MemVariables.resourceData)}")
    print("\texecuting count of domains for each resource")
    ws1 = MemVariables.workbook.create_sheet()
    ws1.title = "resources used"
    ws1.append(("Resource Name", "Count of Domains"))
    ws1.column_dimensions["A"].width = 50
    for theresource, countdomains in sorted(
            MemVariables.resourceData.items(), key=lambda x: x[0].lower()
    ):
        # print(f"\tresource {theresource}:{countdomains}")
        ws1.append((theresource, countdomains))

    # for each resource - count the instances of all domains
    # requires 1 api call per resource
    # print(f"\tgetting resource specific counts resources={len(MemVariables.resourceData)}")
    querystring = {
        "q": "+com.infa.ldm.profiling.dataDomainsAll:*",
        "offset": "0",
        "pageSize": "1",
        "hl": "false",
        "related": "false",
        "rootto": "false",
        "facet.field": [
            "core.classType",
            "com.infa.ldm.profiling.dataDomainsInferred",
            "com.infa.ldm.profiling.dataDomainsAll",
            "com.infa.ldm.profiling.dataDomainsRejected",
            "com.infa.ldm.profiling.dataDomainsAccepted",
        ],
        # "includeRefObjects": "false",
    }

    for resourceName in sorted(MemVariables.resourceData, key=str.casefold):
        print(f"\t{resourceName}", end="")
        querystring["q"] = (
            f"+com.infa.ldm.profiling.dataDomainsAll:* "
            f'+core.resource_name:"{resourceName}"'
        )
        rc, resourceJson = search_summary_v1(MemVariables.edcSession.session, resturl, querystring)
        resitems = resourceJson["totalCount"]
        print(f" objects={resitems:,}")

        resData = dict()
        resClass = dict()
        resRes = dict()
        getDomainQueryStats(resourceJson, resData, resClass, resRes)
        if args.writeToCSV:
            printDomainSummaryToCsv(f"out/domainSummary_{resourceName}.csv", resData)
        printDomainSumarytoExcelWorksheet(MemVariables.workbook, resourceName, resData)

    # save the excel workbook
    print(f"saving output to {MemVariables.outputExcelFile}")
    MemVariables.workbook.save(MemVariables.outputExcelFile)

    print("")
    print(f"Finished - run time = {time.time() - start_time:.2f} seconds")
    print(f"         domains in use={len(MemVariables.domainData)}")
    print(f"items with domain usage={itemCount:,}")
    print(f" resources with domains={len(MemVariables.resourceData)}")

    return


# call main - if not already called or used by another script
if __name__ == "__main__":
    main()
